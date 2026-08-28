#!/usr/bin/env python3
"""Geocode missing shuttle stops and write coordinates back to the Excel workbook.

The workbook is the source of truth. This script only sends rows whose Latitude/
Longitude are blank. Successful results are cached permanently by writing them
back into the workbook, so the same address is not queried again on later runs.

Default provider: OpenStreetMap Foundation's public Nominatim instance.
The endpoint can be replaced without changing code by setting
SHUTTLE_GEOCODER_URL.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少 openpyxl。請先執行：python -m pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "data" / "shuttle-data.xlsx"
DEFAULT_URL = "https://nominatim.openstreetmap.org/search"
GEOCODER_URL = os.environ.get("SHUTTLE_GEOCODER_URL", DEFAULT_URL).strip()
CONTACT = os.environ.get("SHUTTLE_GEOCODER_CONTACT", "").strip()
USER_AGENT = "ShuttleMapGeocoder/1.0" + (f" ({CONTACT})" if CONTACT else "")
MIN_INTERVAL_SECONDS = 1.10

REQUIRED_HEADERS = ["啟用", "StopID", "顯示名稱", "定位搜尋文字", "緯度", "經度"]
STATUS_HEADER = "定位狀態"
TIME_HEADER = "最後定位時間"


def enabled(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or value == "":
        return True
    return str(value).strip().lower() not in {"false", "0", "否", "n", "no"}


def has_number(value) -> bool:
    if value in (None, ""):
        return False
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def geocode(query: str) -> tuple[float, float, str] | None:
    params = urllib.parse.urlencode({
        "format": "jsonv2",
        "limit": 1,
        "countrycodes": "tw",
        "accept-language": "zh-TW",
        "q": query,
    })
    url = f"{GEOCODER_URL}?{params}"
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=20) as response:
        payload = json.loads(response.read().decode("utf-8"))
    if not payload:
        return None
    first = payload[0]
    return float(first["lat"]), float(first["lon"]), str(first.get("display_name", ""))


def ensure_header(ws, header_map: dict[str, int], name: str) -> int:
    if name in header_map:
        return header_map[name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=name)
    header_map[name] = col
    return col


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="自動定位 Excel 中尚未有座標的交通車站點")
    parser.add_argument("--dry-run", action="store_true", help="只列出待定位站點，不連線、不修改 Excel")
    parser.add_argument("--limit", type=int, default=0, help="本次最多查詢幾個站點；0 表示不限")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not XLSX_PATH.exists():
        print(f"找不到 {XLSX_PATH}", file=sys.stderr)
        return 1

    wb = load_workbook(XLSX_PATH)
    if "Stops" not in wb.sheetnames:
        print("Excel 缺少 Stops 工作表", file=sys.stderr)
        return 1
    ws = wb["Stops"]

    header_map = {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value not in (None, "")
    }
    missing = [name for name in REQUIRED_HEADERS if name not in header_map]
    if missing:
        print("Stops 缺少欄位：" + "、".join(missing), file=sys.stderr)
        return 1

    status_col = ensure_header(ws, header_map, STATUS_HEADER)
    time_col = ensure_header(ws, header_map, TIME_HEADER)
    lat_col = header_map["緯度"]
    lng_col = header_map["經度"]
    query_col = header_map["定位搜尋文字"]
    id_col = header_map["StopID"]
    name_col = header_map["顯示名稱"]
    enabled_col = header_map["啟用"]

    pending: list[int] = []
    for row in range(2, ws.max_row + 1):
        if not enabled(ws.cell(row, enabled_col).value):
            continue
        lat = ws.cell(row, lat_col).value
        lng = ws.cell(row, lng_col).value
        if has_number(lat) and has_number(lng):
            if not ws.cell(row, status_col).value:
                ws.cell(row, status_col, "已定位")
            continue
        pending.append(row)

    if args.limit > 0:
        pending = pending[: args.limit]

    print(f"Excel：{XLSX_PATH.relative_to(ROOT)}")
    print(f"待定位：{len(pending)} 個站點")
    if args.dry_run:
        for row in pending:
            sid = str(ws.cell(row, id_col).value or "").strip()
            name = str(ws.cell(row, name_col).value or "").strip()
            query = str(ws.cell(row, query_col).value or "").strip()
            print(f"- {sid} {name} -> {query or '[缺少定位搜尋文字]'}")
        return 0

    success = 0
    failed = 0
    queried = 0
    last_request_at = 0.0

    for idx, row in enumerate(pending, start=1):
        sid = str(ws.cell(row, id_col).value or "").strip()
        name = str(ws.cell(row, name_col).value or "").strip()
        query = str(ws.cell(row, query_col).value or "").strip()
        label = f"{sid} {name}".strip()

        if not query:
            failed += 1
            ws.cell(row, status_col, "失敗：缺少定位搜尋文字")
            ws.cell(row, time_col, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            print(f"✗ {label}：缺少定位搜尋文字")
            continue

        # Respect the public Nominatim absolute maximum of one request per second.
        wait = MIN_INTERVAL_SECONDS - (time.monotonic() - last_request_at)
        if wait > 0:
            time.sleep(wait)

        try:
            queried += 1
            result = geocode(query)
            last_request_at = time.monotonic()
            if result is None:
                failed += 1
                ws.cell(row, status_col, "失敗：找不到結果")
                ws.cell(row, time_col, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                print(f"✗ {label}：找不到結果｜{query}")
            else:
                lat, lng, _display_name = result
                ws.cell(row, lat_col, lat)
                ws.cell(row, lng_col, lng)
                ws.cell(row, status_col, "已定位")
                ws.cell(row, time_col, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                success += 1
                print(f"✓ {label}：{lat:.6f}, {lng:.6f}")
        except urllib.error.HTTPError as exc:
            last_request_at = time.monotonic()
            failed += 1
            ws.cell(row, status_col, f"失敗：HTTP {exc.code}")
            ws.cell(row, time_col, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            print(f"✗ {label}：HTTP {exc.code}｜{query}")
            if exc.code in {403, 429}:
                print("定位服務拒絕或限流，本次停止後續查詢。", file=sys.stderr)
                break
        except (urllib.error.URLError, TimeoutError, OSError, ValueError) as exc:
            last_request_at = time.monotonic()
            failed += 1
            ws.cell(row, status_col, "失敗：網路或回應錯誤")
            ws.cell(row, time_col, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            print(f"✗ {label}：{exc}")
            # A network-wide failure is unlikely to improve for subsequent rows.
            if isinstance(exc, (urllib.error.URLError, TimeoutError)):
                print("偵測到網路連線問題，本次停止後續查詢。", file=sys.stderr)
                break

        # Persist progress periodically so successful geocodes are cached even if interrupted later.
        if queried and queried % 10 == 0:
            wb.save(XLSX_PATH)

    wb.save(XLSX_PATH)
    print()
    print(f"定位完成：成功 {success}、失敗 {failed}、實際送出查詢 {queried}")
    print("成功座標已寫回 Excel；下次不會重複查詢這些站點。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
